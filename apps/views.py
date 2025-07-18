from django.shortcuts import render, redirect
from .forms import ComercioForm
from .models import Comercio
from .recommender import RecomendadorEmpresas
from django.contrib import messages 
import os
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) # Ruta base
PATH_EXCEL = os.path.join(BASE_DIR, '../../data/base_actualizada.xlsx') # Ruta al archivo Excel
recomendador = RecomendadorEmpresas(PATH_EXCEL) # Inicializar el recomendador

def index(request): # Vista principal
    recomendaciones = None
    if request.method == 'POST':
        consulta = request.POST.get('consulta')

        # Recargar recomendador con la base actualizada
        recomendador_actualizado = RecomendadorEmpresas(PATH_EXCEL)
        recomendaciones = recomendador_actualizado.recomendar(consulta).to_dict(orient='records')

    return render(request, 'index.html', {'recomendaciones': recomendaciones})

def registrar_comercio(request): # Vista para registrar un comercio

    if request.method == 'POST': # Manejo del formulario
        form = ComercioForm(request.POST, request.FILES) # Instancia del formulario
        if form.is_valid(): # Validación del formulario
            comercio = form.save() # Guardar el comercio en la base de datos

            # Ruta al archivo Excel
            path_excel = os.path.join(BASE_DIR, '../../data/base_actualizada.xlsx')

            # Verifica si el archivo existe
            if not os.path.exists(path_excel):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append([
                    'ID', 'Nombre', 'Sector', 'Subsector', 'Artículos',
                    'Dirección', 'Celular', 'Teléfono',
                    'Link Maps', 'Facebook', 'Instagram'
                ])
            else:
                wb = openpyxl.load_workbook(path_excel)
                if 'Base de Datos' in wb.sheetnames:    
                    ws = wb['BBDD']
                else:
                    ws = wb.active  # Selecciona la hoja activa si no existe 'Base de Datos'

            # Añade la nueva fila
            ws.append([
                '',  # NIT (vacío por ahora)
                comercio.nombre,
                comercio.sector,
                comercio.subsector,
                comercio.articulos,
                'Colombia',         # País por defecto
                'Antioquia',        # Departamento por defecto
                'Sabaneta',         # Ciudad por defecto
                comercio.direccion,
                comercio.celular,
                comercio.telefono,
                comercio.link_facebook,
                comercio.link_instagram,
            ])

            # Guarda los cambios en el archivo
            wb.save(path_excel)

            messages.success(request, '¡Empresa creada con éxito!')
            return redirect('registro')  
        else:
            messages.error(request, 'Hubo un error al registrar la empresa. Verifica los campos.')
    else:
        form = ComercioForm()
    return render(request, 'registro.html', {'form': form})
